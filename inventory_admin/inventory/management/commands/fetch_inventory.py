import os
import imaplib
import email
import logging
import requests
import hashlib
from io import BytesIO
from openpyxl import load_workbook
from django.conf import settings
from django.core.management.base import BaseCommand

logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = "Fetch inventory Excel file from email and upload only if it contains new rows."

    def handle(self, *args, **kwargs):
        from datetime import datetime
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Inventory fetch started.")

        try:
            creds = self._load_email_credentials()
            msg = self._fetch_email(creds["host"], creds["user"], creds["pass"], "Daily Inventory")
            file_bytes, filename = self._extract_excel_attachment(msg)

            if not file_bytes:
                print("No attachment data found.")
                return

            if not self._has_new_rows(file_bytes):
                print("No new rows found. Skipping upload.")
                return

            self._upload_excel_file(file_bytes, filename)

        except Exception as e:
            logger.exception("Inventory fetch failed:")
            print(f"Inventory fetch failed: {e}")

        print("Inventory fetch completed.")

    def _load_email_credentials(self):
        return {
            "host": "imap.gmail.com",
            "user": "divinexxx07@gmail.com",
            "pass": "djpvsigxtjwchfvo"
        }

    def _fetch_email(self, host, user, password, subject):
        mail = imaplib.IMAP4_SSL(host)
        mail.login(user, password)
        mail.select("inbox")
        _, data = mail.search(None, f'(SUBJECT "{subject}")')
        ids = data[0].split()
        if not ids:
            raise ValueError(f"No email found with subject '{subject}'.")
        latest_id = ids[-1]
        _, msg_data = mail.fetch(latest_id, "(RFC822)")
        return email.message_from_bytes(msg_data[0][1])

    def _extract_excel_attachment(self, msg):
        for part in msg.walk():
            if part.get_content_disposition() == "attachment" and part.get_filename().endswith(".xlsx"):
                return part.get_payload(decode=True), part.get_filename()
        raise ValueError("No .xlsx attachment found in the email.")

    def _hash_row(self, row):
        return hashlib.sha256(",".join(
            str(cell).strip() if cell is not None else "" for cell in row
        ).encode("utf-8")).hexdigest()

    def _has_new_rows(self, file_bytes):
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True)
        except Exception as e:
            print(f"Failed to open Excel file: {e}")
            return False

        sheet_name = "Inventory carrying cost"
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            return False

        sheet = wb[sheet_name]

        cache_path = os.path.join(settings.BASE_DIR, 'uploaded_row_hashes.txt')
        known_hashes = set()
        if os.path.exists(cache_path):
            with open(cache_path, 'r') as f:
                known_hashes = set(f.read().splitlines())

        new_found = False
        new_hashes = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_hash = self._hash_row(row)
            if row_hash not in known_hashes:
                new_found = True
                new_hashes.append(row_hash)

        if new_found:
            with open(cache_path, "a") as f:
                for h in new_hashes:
                    f.write(h + "\n")
            print(f"{len(new_hashes)} new rows detected and tracked.")

        return new_found

    def _upload_excel_file(self, file_bytes, filename):
        print("Uploading Excel file ...")
        url = "http://127.0.0.1:8000/carrying-cost-upload/"
        files = {
            "file": (
                filename,
                BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        try:
            response = requests.post(url, files=files, timeout=60)
            if response.status_code in [200, 201]:
                print("Upload succeeded.")
            else:
                print(f"Upload failed – {response.status_code}: {response.text}")
        except Exception as e:
            print(f"Upload exception: {e}")
